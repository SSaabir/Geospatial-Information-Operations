"""
Reports API - PDF and Excel Export
Provides report generation and export functionality for Researcher and Professional tiers.
Counts as AI operations toward quota.
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from typing import Literal, Optional
from datetime import datetime, timedelta
from pathlib import Path
import io
import json

# PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference

from fastapi.responses import StreamingResponse
from security.auth_middleware import get_current_user
from models.user import UserDB
from models.usage import UsageMetrics
from db_config import DatabaseConfig
from utils.tier import check_and_notify_usage, enforce_quota_or_raise
import pandas as pd

reports_router = APIRouter(prefix="/api/reports", tags=["Reports"])

# Database configuration
db_config = DatabaseConfig()

def get_db():
    """Database session dependency"""
    db = db_config.get_session()
    try:
        yield db
    finally:
        db.close()

# Schema definitions
from pydantic import BaseModel, Field

class ReportRequest(BaseModel):
    """Report generation request"""
    format: Literal["pdf", "excel"] = Field(..., description="Export format")
    report_type: str = Field(..., description="Type of report to generate")
    days_back: int = Field(default=7, ge=1, le=90, description="Number of days of historical data to include")
    include_charts: bool = Field(True, description="Include charts and visualizations")
    workflow_data: Optional[dict] = Field(None, description="Workflow output data from chat")


def generate_pdf_report(
    report_type: str,
    data: dict,
    username: str,
    tier: str,
    include_charts: bool = True,
    workflow_data: Optional[dict] = None
) -> io.BytesIO:
    """
    Generate a comprehensive PDF report using ReportLab with workflow data
    
    Args:
        report_type: Type of report 
        data: Report data dictionary (legacy)
        username: User's username
        tier: User's subscription tier
        include_charts: Whether to include visual elements
        workflow_data: Actual workflow output from chat message
        
    Returns:
        BytesIO buffer containing PDF
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    subheading_style = ParagraphStyle(
        'CustomSubHeading',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=colors.HexColor('#4b5563'),
        spaceAfter=8,
        spaceBefore=8,
        fontName='Helvetica-Bold'
    )
    
    body_style = ParagraphStyle(
        'CustomBody',
        parent=styles['BodyText'],
        fontSize=10,
        textColor=colors.black,
        spaceAfter=6,
        leading=14
    )
    
    # Header
    story.append(Paragraph(f"{report_type.replace('_', ' ').title()}", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Metadata table
    metadata = [
        ['Generated for:', username],
        ['Subscription Tier:', tier.upper()],
        ['Generated on:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Report Type:', report_type.replace('_', ' ').title()],
    ]
    
    meta_table = Table(metadata, colWidths=[2*inch, 4*inch])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Use workflow_data if available
    if workflow_data:
        # Extract data table from collector_data
        collector_data = workflow_data.get('collector_data')
        if collector_data:
            story.append(Paragraph("Data Table", heading_style))
            
            # Parse collector_data if it's a string
            if isinstance(collector_data, str):
                try:
                    collector_data = json.loads(collector_data)
                except:
                    pass
            
            # Extract rows
            rows = None
            if isinstance(collector_data, dict):
                data_obj = collector_data.get('data', {})
                if isinstance(data_obj, dict):
                    rows = data_obj.get('rows', [])
                elif isinstance(data_obj, list):
                    rows = data_obj
            
            if rows and len(rows) > 0:
                # Build table with headers
                headers = list(rows[0].keys()) if rows else []
                table_data = [headers]
                for row in rows[:50]:  # Limit to 50 rows
                    table_data.append([str(row.get(h, '')) for h in headers])
                
                # Calculate column widths
                num_cols = len(headers)
                col_width = 6.5 * inch / num_cols
                
                data_table = Table(table_data, colWidths=[col_width] * num_cols)
                data_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
                ]))
                story.append(data_table)
                story.append(Spacer(1, 0.3*inch))
        
        # Add report content
        report_content = workflow_data.get('report_content', '')
        if report_content:
            story.append(Paragraph("Report Analysis", heading_style))
            
            # Parse sections
            sections = report_content.split('\n\n')
            for section in sections:
                if section.strip():
                    lines = section.strip().split('\n')
                    # First line as subheading if it's short and uppercase-ish
                    if lines and len(lines[0]) < 50 and any(c.isupper() for c in lines[0]):
                        story.append(Paragraph(lines[0], subheading_style))
                        body_lines = lines[1:]
                    else:
                        body_lines = lines
                    
                    # Body text
                    for line in body_lines:
                        if line.strip().startswith('-'):
                            story.append(Paragraph(f"• {line.strip()[1:].strip()}", body_style))
                        elif line.strip():
                            story.append(Paragraph(line.strip(), body_style))
            
            story.append(Spacer(1, 0.3*inch))
        
        # Add visualizations
        visualizations = workflow_data.get('visualizations', [])
        print(f"[DEBUG] Visualizations type: {type(visualizations)}, value: {visualizations}")
        
        # Convert dict to list if needed (orchestrator returns dict)
        if isinstance(visualizations, dict):
            visualizations = list(visualizations.values())
        elif not isinstance(visualizations, list):
            visualizations = []
        
        if visualizations and include_charts:
            story.append(Paragraph("Visualizations", heading_style))
            
            images_added = 0
            for viz_path in visualizations[:4]:  # Limit to 4 images
                try:
                    # Skip None or empty paths
                    if not viz_path:
                        continue
                    
                    # Check if it's a URL or local path
                    if isinstance(viz_path, str) and viz_path.startswith('http'):
                        # Download image
                        import urllib.request
                        img_data = urllib.request.urlopen(viz_path).read()
                        img_buffer = io.BytesIO(img_data)
                        img = Image(img_buffer, width=5*inch, height=3*inch)
                    elif isinstance(viz_path, str) and Path(viz_path).exists():
                        # Local file - verify it exists
                        img = Image(viz_path, width=5*inch, height=3*inch)
                    else:
                        print(f"Skipping invalid or non-existent image path: {viz_path}")
                        continue
                    
                    story.append(img)
                    story.append(Spacer(1, 0.2*inch))
                    images_added += 1
                except Exception as e:
                    print(f"Error loading image {viz_path}: {e}")
                    # Continue processing other images
                    continue
            
            # If no images were added, remove the "Visualizations" heading
            if images_added == 0:
                # Remove last two items (heading and spacer before it)
                story = story[:-1]
    
    # Fallback to old report format if no workflow_data
    elif report_type == "weather_summary":
        story.append(Paragraph("Weather Summary", heading_style))
        
        # Summary statistics
        summary_data = data.get('summary', {})
        summary_items = [
            ['Metric', 'Value'],
            ['Average Temperature', f"{summary_data.get('avg_temp', 0):.1f}°C"],
            ['Max Temperature', f"{summary_data.get('max_temp', 0):.1f}°C"],
            ['Min Temperature', f"{summary_data.get('min_temp', 0):.1f}°C"],
            ['Average Humidity', f"{summary_data.get('avg_humidity', 0):.1f}%"],
            ['Total Rainfall', f"{summary_data.get('total_rain', 0):.1f}mm"],
        ]
        
        summary_table = Table(summary_items, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        
    elif report_type == "forecast_analysis":
        story.append(Paragraph("Forecast Analysis", heading_style))
        story.append(Paragraph(
            "This report provides detailed forecast predictions based on historical trends and AI models.",
            styles['Normal']
        ))
        story.append(Spacer(1, 0.2*inch))
        
        # Forecast data
        forecast_data = data.get('forecast', [])
        if forecast_data:
            forecast_items = [['Date', 'Temperature', 'Humidity', 'Conditions', 'Confidence']]
            for item in forecast_data[:10]:  # Limit to 10 entries
                forecast_items.append([
                    item.get('date', 'N/A'),
                    f"{item.get('temperature', 0):.1f}°C",
                    f"{item.get('humidity', 0):.0f}%",
                    item.get('conditions', 'N/A'),
                    f"{item.get('confidence', 0):.0%}"
                ])
            
            forecast_table = Table(forecast_items, colWidths=[1.5*inch, 1.2*inch, 1.2*inch, 1.8*inch, 1.3*inch])
            forecast_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            story.append(forecast_table)
    
    elif report_type == "usage_stats":
        story.append(Paragraph("Usage Statistics", heading_style))
        
        usage_data = data.get('usage', {})
        usage_items = [
            ['Metric', 'Count', 'Limit'],
            ['AI Operations', str(usage_data.get('api_calls', 0)), str(usage_data.get('limit', 'Unlimited'))],
            ['Reports Generated', str(usage_data.get('reports_generated', 0)), 'N/A'],
            ['Usage Percentage', f"{usage_data.get('usage_percent', 0):.1f}%", '100%'],
        ]
        
        usage_table = Table(usage_items, colWidths=[2.5*inch, 1.75*inch, 1.75*inch])
        usage_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(usage_table)
    
    # Footer
    story.append(Spacer(1, 0.5*inch))
    footer_text = Paragraph(
        f"<i>This report was generated by Geospatial Information Operations Platform. "
        f"For questions, contact support@geo-ops.com</i>",
        styles['Italic']
    )
    story.append(footer_text)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_excel_report(
    report_type: str,
    data: dict,
    username: str,
    tier: str,
    include_charts: bool = True
) -> io.BytesIO:
    """
    Generate an Excel report using openpyxl
    
    Args:
        report_type: Type of report
        data: Report data dictionary
        username: User's username
        tier: User's subscription tier
        include_charts: Whether to include charts
        
    Returns:
        BytesIO buffer containing Excel file
    """
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = report_type.replace('_', ' ').title()
    
    # Styling
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = report_type.replace('_', ' ').title()
    ws['A1'].font = Font(size=18, bold=True, color="1E40AF")
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Metadata
    row = 3
    metadata = [
        ('Generated for:', username),
        ('Tier:', tier.upper()),
        ('Generated on:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('Period:', f"Last {data.get('days_back', 7)} days")
    ]
    
    for label, value in metadata:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    row += 1  # Spacer
    
    # Report-specific data
    if report_type == "weather_summary":
        summary = data.get('summary', {})
        
        # Headers
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Data
        metrics = [
            ('Average Temperature', f"{summary.get('avg_temp', 0):.1f}°C"),
            ('Max Temperature', f"{summary.get('max_temp', 0):.1f}°C"),
            ('Min Temperature', f"{summary.get('min_temp', 0):.1f}°C"),
            ('Average Humidity', f"{summary.get('avg_humidity', 0):.1f}%"),
            ('Total Rainfall', f"{summary.get('total_rain', 0):.1f}mm"),
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
            
    elif report_type == "forecast_analysis":
        forecast_data = data.get('forecast', [])
        
        # Headers
        headers = ['Date', 'Temperature', 'Humidity', 'Conditions', 'Confidence']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Data
        for item in forecast_data:
            ws.cell(row=row, column=1, value=item.get('date', 'N/A'))
            ws.cell(row=row, column=2, value=f"{item.get('temperature', 0):.1f}°C")
            ws.cell(row=row, column=3, value=f"{item.get('humidity', 0):.0f}%")
            ws.cell(row=row, column=4, value=item.get('conditions', 'N/A'))
            ws.cell(row=row, column=5, value=f"{item.get('confidence', 0):.0%}")
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
            row += 1
            
    elif report_type == "usage_stats":
        usage = data.get('usage', {})
        
        # Headers
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Count"
        ws[f'C{row}'] = "Limit"
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Data
        stats = [
            ('AI Operations', usage.get('api_calls', 0), usage.get('limit', 'Unlimited')),
            ('Reports Generated', usage.get('reports_generated', 0), 'N/A'),
            ('Usage Percentage', f"{usage.get('usage_percent', 0):.1f}%", '100%'),
        ]
        
        for metric, count, limit in stats:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = str(count)
            ws[f'C{row}'] = str(limit)
            for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
                cell.border = border
            row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@reports_router.post("/export")
async def export_report(
    request: ReportRequest,
    current_user: UserDB = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Generate and export a report in PDF or Excel format.
    
    **Tier Requirements:** Researcher or Professional tier only
    **Quota:** Counts as 1 AI operation
    
    Args:
        request: Report generation parameters
        current_user: Authenticated user
        db: Database session
        
    Returns:
        StreamingResponse with PDF or Excel file
        
    Raises:
        HTTPException 403: If user is on Free tier
        HTTPException 402: If quota exceeded
    """
    # Get user tier and metrics
    user_tier = getattr(current_user, "tier", "free")
    username = current_user.username or current_user.email
    
    # Tier check - Only Researcher and Professional can generate reports
    if user_tier == 'free':
        raise HTTPException(
            status_code=403,
            detail="PDF/Excel reports require Researcher or Professional tier. Upgrade to access this feature."
        )
    
    # Get usage metrics
    metrics = db.query(UsageMetrics).filter(UsageMetrics.user_id == current_user.id).first()
    if not metrics:
        metrics = UsageMetrics(user_id=current_user.id)
        db.add(metrics)
        db.commit()
        db.refresh(metrics)
    
    # Check quota and send notifications
    check_and_notify_usage(metrics, user_tier, current_user.id, username)
    
    # Enforce quota (raises exception if exceeded)
    enforce_quota_or_raise(metrics, user_tier, current_user.id, username)
    
    # Prepare report data (mock data for now - can be replaced with actual data queries)
    report_data = {
        'days_back': request.days_back,
        'summary': {
            'avg_temp': 28.5,
            'max_temp': 32.1,
            'min_temp': 24.3,
            'avg_humidity': 75.2,
            'total_rain': 45.3
        },
        'forecast': [
            {
                'date': (datetime.now() + timedelta(days=i)).strftime('%Y-%m-%d'),
                'temperature': 28.0 + i * 0.5,
                'humidity': 70 + i * 2,
                'conditions': 'Partly Cloudy',
                'confidence': 0.85 - (i * 0.02)
            }
            for i in range(request.days_back)
        ],
        'usage': {
            'api_calls': metrics.api_calls,
            'reports_generated': metrics.reports_generated,
            'limit': 5000 if user_tier == 'researcher' else float('inf'),
            'usage_percent': (metrics.api_calls / 5000 * 100) if user_tier == 'researcher' else 0
        }
    }
    
    # Generate report based on format
    try:
        print(f"[DEBUG] Generating PDF with workflow_data keys: {request.workflow_data.keys() if request.workflow_data else 'None'}")
        if request.workflow_data:
            print(f"[DEBUG] Workflow type: {request.workflow_data.get('workflow_type')}")
            print(f"[DEBUG] Has visualizations: {'visualizations' in request.workflow_data}")
        
        if request.format == "pdf":
            buffer = generate_pdf_report(
                request.report_type,
                report_data,
                username,
                user_tier,
                request.include_charts,
                workflow_data=request.workflow_data
            )
            media_type = "application/pdf"
            filename = f"{request.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        else:  # excel
            buffer = generate_excel_report(
                request.report_type,
                report_data,
                username,
                user_tier,
                request.include_charts
            )
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"{request.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Increment counters
        metrics.api_calls += 1
        metrics.reports_generated += 1
        db.commit()
    except Exception as e:
        # Log the error and return detailed message
        import traceback
        error_details = traceback.format_exc()
        print(f"Report generation error: {error_details}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to generate report: {str(e)}"
        )
    
    # Return as streaming response
    return StreamingResponse(
        buffer,
        media_type=media_type,
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
